#!/usr/bin/env python3
"""
aggiorna_portale.py
-------------------
Legge il file Excel di IncoService da OneDrive e aggiorna
automaticamente il file portale_incoservice.html con i dati aggiornati.

Poi esegue git add + commit + push su GitHub Pages.

Configurazione: modifica le variabili nella sezione CONFIG qui sotto.
"""

import os
import re
import sys
import json
import time
import shutil
import datetime
import subprocess
import openpyxl

# ============================================================
# CONFIG — modifica questi percorsi
# ============================================================

# Percorso completo al file Excel su OneDrive
# Esempio Windows: r""C:\Users\Davide\OneDrive - INCOSERVICE SRL\File di Microsoft Copilot Chat\GestionaleIncoservice\011 - Programmazione settimanale LAVORAZIONI\Gestione Produzioni_rev.xlsx""
EXCEL_PATH = r""C:\Users\Davide\OneDrive - INCOSERVICE SRL\File di Microsoft Copilot Chat\GestionaleIncoservice\011 - Programmazione settimanale LAVORAZIONI\Gestione Produzioni_rev.xlsx""

# Percorso alla cartella del repository GitHub locale
# Esempio: r"C:\Users\TuoNome\Documents\portale-incoservice"
REPO_PATH = r""C:\Users\Davide\OneDrive - INCOSERVICE SRL\Attachments\Desktop\portale-incoservice""

# Nome del file HTML nel repository
HTML_FILENAME = "index.html"

# Quanti minuti tra un aggiornamento e l'altro (5 = ogni 5 minuti)
INTERVALLO_MINUTI = 5

# ============================================================
# FINE CONFIG
# ============================================================

HTML_PATH = os.path.join(REPO_PATH, HTML_FILENAME)


def log(msg):
    ts = datetime.datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}")


def get_row_color(ws, row):
    """Determina il colore della riga: gialla / colorata / bianca"""
    cell = ws.cell(row, 1)
    fill = cell.fill
    try:
        if fill.fgColor.type == "rgb":
            rgb = fill.fgColor.rgb
            if rgb == "FFFFFF00":
                return "gialla"
            elif rgb in ("00000000", "00FFFFFF", "FFFFFFFF") or not rgb:
                return "bianca"
            else:
                return "colorata"
        elif fill.fgColor.type == "theme":
            return "colorata"
        else:
            return "bianca"
    except Exception:
        return "bianca"


def fmt_date(val):
    """Converte datetime Excel in stringa YYYY-MM-DD, o restituisce il valore così com'è."""
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, datetime.date):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    return s if s else None


def cell_val(ws, row, col):
    """Legge una cella e restituisce il valore pulito (None se vuoto)."""
    v = ws.cell(row, col).value
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def js_val(v):
    """Converte un valore Python in stringa JavaScript."""
    if v is None:
        return "null"
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, (int, float)):
        return str(v)
    # Stringa: escape apici
    s = str(v).replace("\\", "\\\\").replace('"', '\\"')
    return f'"{s}"'


def leggi_ordini(ws):
    """Legge il foglio Programma Produzioni e restituisce lista di dict."""
    ordini = []
    for row in range(2, ws.max_row + 1):
        cliente = cell_val(ws, row, 1)
        if not cliente:
            continue  # riga vuota

        colore = get_row_color(ws, row)
        pronto = colore == "gialla"

        # Leggi ODL (colonne 23-29)
        odl = []
        for c in range(23, 30):
            v = ws.cell(row, c).value
            if v and str(v).strip():
                try:
                    odl.append(int(v))
                except (ValueError, TypeError):
                    pass

        impegno   = cell_val(ws, row, 2)
        tipologia = cell_val(ws, row, 7)
        qty       = cell_val(ws, row, 8)
        finitura  = cell_val(ws, row, 9)
        ral       = cell_val(ws, row, 10)
        posa      = cell_val(ws, row, 11)

        # Fase produzione (col 12) — se "-" = N/A
        produzione    = cell_val(ws, row, 12)
        invioZN       = fmt_date(cell_val(ws, row, 14)) or cell_val(ws, row, 14)
        ritornoZN     = fmt_date(cell_val(ws, row, 16)) or cell_val(ws, row, 16)
        pulizia       = fmt_date(cell_val(ws, row, 17)) or cell_val(ws, row, 17)
        invioRAL      = fmt_date(cell_val(ws, row, 18)) or cell_val(ws, row, 18)
        verniciatura  = cell_val(ws, row, 19)
        imballo       = fmt_date(cell_val(ws, row, 20)) or cell_val(ws, row, 20)
        consStimata   = fmt_date(cell_val(ws, row, 21))
        consRichiesta = fmt_date(cell_val(ws, row, 22))

        # La colonna 13 è vuota (colonna nascosta), col 15 = "zn" (valore interno)
        # Se col 15 ha valore "NZ" o simile = invioZN già completato ma ritornoZN pianificato
        # Già gestito tramite null/valore nel portale

        ordini.append({
            "cliente":      str(cliente).strip(),
            "impegno":      str(impegno) if impegno is not None else "",
            "tipologia":    str(tipologia) if tipologia else "",
            "qty":          str(qty) if qty is not None else "",
            "finitura":     str(finitura).lower() if finitura else None,
            "ral":          str(ral) if ral else None,
            "posa":         str(posa) if posa else None,
            "produzione":   str(produzione) if produzione else None,
            "invioZN":      invioZN,
            "ritornoZN":    ritornoZN,
            "pulizia":      pulizia,
            "invioRAL":     invioRAL,
            "verniciatura": str(verniciatura) if verniciatura else None,
            "imballo":      imballo,
            "consStimata":  consStimata,
            "consRichiesta":consRichiesta,
            "odl":          odl,
            "prontoConsegna": pronto,
            "colore":       colore,
        })

    return ordini


def leggi_pose(ws):
    """Legge il foglio Programma Pose."""
    pose = []
    # Cerca la riga di intestazione (contiene "CLIENTE" o "cliente")
    header_row = None
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row, col).value
            if v and str(v).strip().upper() in ("CLIENTE", "SETTIMANA", "DATA"):
                header_row = row
                break
        if header_row:
            break

    if not header_row:
        return pose  # foglio Pose vuoto o formato diverso

    # Mappa intestazioni
    headers = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(header_row, col).value
        if v:
            headers[str(v).strip().lower()] = col

    for row in range(header_row + 1, ws.max_row + 1):
        cliente = cell_val(ws, row, headers.get("cliente", 1))
        if not cliente:
            continue
        data_raw = ws.cell(row, headers.get("data", 3)).value
        data = fmt_date(data_raw)
        pose.append({
            "settimana":  str(cell_val(ws, row, headers.get("settimana", 1)) or ""),
            "giorno":     str(cell_val(ws, row, headers.get("giorno", 2)) or ""),
            "data":       data or "",
            "cliente":    str(cliente),
            "impegno":    str(cell_val(ws, row, headers.get("impegno", 4)) or "-"),
            "luogo":      str(cell_val(ws, row, headers.get("luogo", 5)) or ""),
            "note":       str(cell_val(ws, row, headers.get("note", 6)) or ""),
            "confermato": cell_val(ws, row, headers.get("confermato", 7)),
        })
    return pose


def genera_js_ordini(ordini):
    """Genera il blocco JavaScript const ORDINI_DATA = [...]"""
    lines = ["const ORDINI_DATA = ["]
    for o in ordini:
        odl_js = "[" + ",".join(str(x) for x in o["odl"]) + "]"
        line = (
            f'  {{ cliente:{js_val(o["cliente"])}, impegno:{js_val(o["impegno"])}, '
            f'tipologia:{js_val(o["tipologia"])}, qty:{js_val(o["qty"])}, '
            f'finitura:{js_val(o["finitura"])}, ral:{js_val(o["ral"])}, '
            f'posa:{js_val(o["posa"])}, '
            f'produzione:{js_val(o["produzione"])}, invioZN:{js_val(o["invioZN"])}, '
            f'ritornoZN:{js_val(o["ritornoZN"])}, pulizia:{js_val(o["pulizia"])}, '
            f'invioRAL:{js_val(o["invioRAL"])}, verniciatura:{js_val(o["verniciatura"])}, '
            f'imballo:{js_val(o["imballo"])}, '
            f'consStimata:{js_val(o["consStimata"])}, consRichiesta:{js_val(o["consRichiesta"])}, '
            f'odl:{odl_js}, prontoConsegna:{js_val(o["prontoConsegna"])}, '
            f'colore:{js_val(o["colore"])} }},'
        )
        lines.append(line)
    lines.append("];")
    return "\n".join(lines)


def genera_js_pose(pose):
    """Genera il blocco JavaScript const POSE_DATA = [...]"""
    lines = ["const POSE_DATA = ["]
    for p in pose:
        line = (
            f'  {{ settimana:{js_val(p["settimana"])}, giorno:{js_val(p["giorno"])}, '
            f'data:{js_val(p["data"])}, cliente:{js_val(p["cliente"])}, '
            f'impegno:{js_val(p["impegno"])}, luogo:{js_val(p["luogo"])}, '
            f'note:{js_val(p["note"])}, confermato:{js_val(p["confermato"])} }},'
        )
        lines.append(line)
    lines.append("];")
    return "\n".join(lines)


def aggiorna_html(ordini, pose):
    """Sostituisce ORDINI_DATA e POSE_DATA nell'HTML."""
    with open(HTML_PATH, "r", encoding="utf-8") as f:
        html = f.read()

    # Sostituisce ORDINI_DATA
    pattern_ordini = r"const ORDINI_DATA\s*=\s*\[.*?\];"
    nuovo_ordini = genera_js_ordini(ordini)
    html_new, n1 = re.subn(pattern_ordini, nuovo_ordini, html, flags=re.DOTALL)
    if n1 == 0:
        log("ERRORE: ORDINI_DATA non trovato nell'HTML!")
        return False

    # Sostituisce POSE_DATA
    pattern_pose = r"const POSE_DATA\s*=\s*\[.*?\];"
    nuovo_pose = genera_js_pose(pose)
    html_new, n2 = re.subn(pattern_pose, nuovo_pose, html_new, flags=re.DOTALL)
    if n2 == 0:
        log("ATTENZIONE: POSE_DATA non trovato nell'HTML, saltato.")

    with open(HTML_PATH, "w", encoding="utf-8") as f:
        f.write(html_new)

    log(f"HTML aggiornato: {len(ordini)} ordini, {len(pose)} pose.")
    return True


def git_push():
    """Esegue git add, commit e push."""
    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    try:
        subprocess.run(["git", "-C", REPO_PATH, "add", HTML_FILENAME],
                       check=True, capture_output=True)
        result = subprocess.run(
            ["git", "-C", REPO_PATH, "diff", "--cached", "--quiet"],
            capture_output=True
        )
        if result.returncode == 0:
            log("Nessuna modifica da pubblicare.")
            return True  # niente da fare

        subprocess.run(
            ["git", "-C", REPO_PATH, "commit", "-m", f"Aggiornamento dati {ts}"],
            check=True, capture_output=True
        )
        subprocess.run(
            ["git", "-C", REPO_PATH, "push"],
            check=True, capture_output=True
        )
        log(f"Pubblicato su GitHub Pages!")
        return True
    except subprocess.CalledProcessError as e:
        log(f"Errore git: {e.stderr.decode() if e.stderr else e}")
        return False


def ciclo():
    """Esegue un ciclo completo: leggi Excel -> aggiorna HTML -> push."""
    log("Lettura Excel...")
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    except Exception as e:
        log(f"Errore apertura Excel: {e}")
        return

    try:
        ws_prod = wb["Programma Produzioni"]
        ordini = leggi_ordini(ws_prod)
        log(f"Letti {len(ordini)} ordini.")
    except Exception as e:
        log(f"Errore lettura Produzioni: {e}")
        return

    try:
        ws_pose = wb["Programma Pose"]
        pose = leggi_pose(ws_pose)
        log(f"Lette {len(pose)} pose.")
    except Exception as e:
        log(f"Foglio Pose non letto: {e}")
        pose = []

    if not aggiorna_html(ordini, pose):
        return

    git_push()


def main():
    log("=== Avvio aggiornamento automatico portale IncoService ===")
    log(f"Excel:    {EXCEL_PATH}")
    log(f"Repo:     {REPO_PATH}")
    log(f"Intervallo: ogni {INTERVALLO_MINUTI} minuti")
    log("Premi Ctrl+C per fermare.")
    print()

    while True:
        try:
            ciclo()
        except Exception as e:
            log(f"Errore imprevisto: {e}")

        log(f"Prossimo aggiornamento tra {INTERVALLO_MINUTI} minuti...")
        time.sleep(INTERVALLO_MINUTI * 60)


if __name__ == "__main__":
    main()
